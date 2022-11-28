import smtplib 
import os
import openpyxl

EMAIL_ADDRESS = os.environ.get('EMAIL_ADDRESS')
EMAIL_PASSWORD = os.environ.get('CHROME_TOKEN')

companies = ["datadog", "microsoft","google","salesforce","td","shopify","spotify","amazon","rbc","scotiabank","qualcomm","robinhood","grabyo","siltronic","cisco","hashicorp","genentech","qualtrics","linkedin","paypal","docusign", "bloomberg", "lyft", "databricks","tangerine","square", "nintendo", "snap", "league", "wealthsimple", "uber", "bmo"]
formats = {"name":"", "company":""}

with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
    smtp.ehlo()
    smtp.starttls()
    smtp.ehlo()

    smtp.login(EMAIL_ADDRESS, "upow nxwo fayd dacw") 

    body = "Hi {0[name]},\n\nI hope this email finds you well. My name is Nick, and I am currently in my third year at Ivey business school, completing a dual degree with computer science. I'm interested in hearing about your experiences at {1[company]}.\n\nIn terms of my background, I spent my first year summer interning at Relay Financial as a data analyst and spent my last summer doing research for professor at Ivey, in the creation of a faculty database of North American universities.\n\nIf you're willing, I would appreciate the opportunity to learn more about your experiences through a call. I am more than happy to work around your schedule.\n\nBest regards,\n\nNick\n\nNicholas Lam\nComputer Science, HBA1\nIvey Business School\nTel: 647-528-2860"

    subject = "Western Student Reaching Out"

    df = openpyxl.load_workbook("CAD_Data.xlsx")
    df = df.active
    for r in range(201,300):
        name = df.cell(row=r, column=1).value
        if name:
            name = name.split(" ")[0].title()
            formats["name"] = name
        else:
            print("No name!")
            continue

        company = df.cell(row=r, column=2).value
        if company:
            clist = company.split(" ")
            found = False

            for c in clist:
                if c.lower() in companies:
                    company = c
                    found = True
                    break
                
            if not found:
                print("Company invalid")
                continue
            
            if company == "bmo" or company == "RBC" or company == "TD":
                company = company.upper()
            else:
                company = company.title()

            formats["company"] = company
        else:
            print("No Company!")
            continue

        email = df.cell(row=r, column=4).value
        if email == "N/A" or email == "['Company not recognized!']":
            print("Invalid Email!")
            continue

        emails = email.replace("'", "").replace("[","").replace("]","").split(",")
        body = "Hi {0[name]},\n\nI hope this email finds you well. My name is Nick, and I am currently in my third year at Ivey business school, completing a dual degree with computer science. I'm interested in hearing about your experiences at {1[company]}.\n\nIn terms of my background, I spent my first year summer interning at Relay Financial as a data analyst and spent my last summer doing research for a professor at Ivey, in the creation of a faculty database of North American universities.\n\nIf you're willing, I would appreciate the opportunity to learn more about your experiences through a call. I am more than happy to work around your schedule.\n\nBest regards,\n\nNick\n\nNicholas Lam\nComputer Science, HBA1\nIvey Business School\nTel: 647-528-2860".format(formats, formats)

        for e in emails:
            print(e)
            msg = f'Subject: {subject}\n\n{body}'
            if msg is None:
                continue
            try:
                smtp.sendmail("nicholasalexanderlam@gmail.com", e, msg.encode('utf-8'))
            except:
                continue